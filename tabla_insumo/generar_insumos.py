# -*- coding: utf-8 -*-
"""
Generador de insumos COIPO — fuente de verdad de los catálogos compartidos.

Hace dos cosas:
  1) SEED: crea/repuebla los .xlsx de esta carpeta a partir de los datos
     canónicos embebidos abajo (region_comuna_vivero, plantas, stock,
     parametros_region).
  2) BUILD: lee esos .xlsx y emite un único `catalogo.generated.json` en
     AMBOS frontends (ENTREGA e INVENTARIO). Ese JSON es el contrato
     compartido que vuelve interoperables a las dos plataformas.

Uso:
  python generar_insumos.py            # seed (si falta) + build
  python generar_insumos.py --seed     # fuerza re-seed (sobrescribe los .xlsx) + build
  python generar_insumos.py --build    # solo build (lee los .xlsx existentes)

Requiere: openpyxl  (pip install openpyxl) — el mismo que ya usa INSUMO/_extract.py
"""

import os
import re
import sys
import json
import unicodedata

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))                       # .../COIPO_ENTREGA_PLANTA/tabla_insumo
ENTREGA_ROOT = os.path.dirname(BASE)                                    # .../COIPO_ENTREGA_PLANTA
GITHUB_ROOT = os.path.dirname(ENTREGA_ROOT)                             # .../GitHub
INVENTARIO_ROOT = os.path.join(GITHUB_ROOT, "COIPO_INVENTARIO_PLANTA")

# Destinos del catálogo JSON compartido (uno por frontend) + copia de referencia.
JSON_TARGETS = [
    os.path.join(ENTREGA_ROOT, "frontend", "src", "data", "catalogo.generated.json"),
    os.path.join(INVENTARIO_ROOT, "frontend", "src", "data", "catalogo.generated.json"),
    os.path.join(BASE, "catalogo.generated.json"),
]

XLSX = {
    "region_comuna_vivero": os.path.join(BASE, "region_comuna_vivero.xlsx"),
    "plantas": os.path.join(BASE, "plantas.xlsx"),
    "stock": os.path.join(BASE, "stock.xlsx"),
    "parametros_region": os.path.join(BASE, "parametros_region.xlsx"),
}

# ---------------------------------------------------------------------------
# DATOS CANÓNICOS (semilla) — derivados de
# COIPO_ENTREGA_PLANTA/frontend/src/api/mockData.js
# ---------------------------------------------------------------------------

REGIONES = [
    (1, "Arica y Parinacota"),
    (2, "Tarapacá"),
    (3, "Antofagasta"),
    (4, "Atacama"),
    (5, "Coquimbo"),
    (6, "Valparaíso"),
    (7, "Libertador General Bernardo O'Higgins"),
    (8, "Maule"),
    (9, "Ñuble"),
    (10, "Biobío"),
    (11, "La Araucanía"),
    (12, "Los Ríos"),
    (13, "Los Lagos"),
    (14, "Aysén"),
    (15, "Magallanes y La Antártica Chilena"),
    (16, "Metropolitana de Santiago"),
    (17, "Isla de Pascua"),
]

# (vivero_id, nombre, region_id, provincia, comuna, encargado_email)
VIVEROS = [
    (1,  "Vivero Las Maitas",                              1,  "Arica",        "Arica",          "enc.lasmaitas@conaf.cl"),
    (2,  "Vivero Putre",                                   1,  "Parinacota",   "Putre",          "enc.putre@conaf.cl"),
    (3,  "Vivero Alejandro Caipa",                         2,  "Tamarugal",    "Pozo Almonte",   "enc.caipa@conaf.cl"),
    (4,  "Vivero Regional Antofagasta",                    3,  "Antofagasta",  "Antofagasta",    "enc.antofagasta@conaf.cl"),
    (5,  "Vivero Quillagua",                               3,  "Tocopilla",    "María Elena",    "enc.quillagua@conaf.cl"),
    (6,  "Vivero Copiapó",                                 4,  "Copiapó",      "Copiapó",        "enc.copiapo@conaf.cl"),
    (7,  "Vivero Vallenar",                                4,  "Huasco",       "Vallenar",       "enc.vallenar@conaf.cl"),
    (8,  "Vivero Illapel",                                 5,  "Choapa",       "Illapel",        "enc.illapel@conaf.cl"),
    (9,  "Vivero Archipiélago Juan Fernández",             6,  "Valparaíso",   "Juan Fernández", "enc.juanfernandez@conaf.cl"),
    (10, "Vivero Reserva Nacional Lago Peñuelas",          6,  "Valparaíso",   "Valparaíso",     "enc.penuelas@conaf.cl"),
    (11, "Vivero La Ligua",                                6,  "Petorca",      "La Ligua",       "enc.laligua@conaf.cl"),
    (12, "Vivero Llay Llay",                               6,  "San Felipe",   "Llay Llay",      "enc.llayllay@conaf.cl"),
    (13, "Vivero Chomedahue",                              7,  "Colchagua",    "Santa Cruz",     "enc.chomedahue@conaf.cl"),
    (14, "Centro de Acopio Pantanillo",                    8,  "Talca",        "Constitución",   "enc.pantanillo@conaf.cl"),
    (15, "Centro Experimental Cauquenes",                  8,  "Cauquenes",    "Cauquenes",      "enc.cauquenes@conaf.cl"),
    (16, "Vivero Centro de Semillas, Genética y Entomología", 9, "Diguillín",  "Chillán",        "enc.chillan@conaf.cl"),
    (17, "Base Brigada La Granja",                         10, "Arauco",       "Cañete",         "enc.lagranja@conaf.cl"),
    (18, "Centro de Acondicionamiento Duqueco",            10, "Biobío",       "Los Ángeles",    "enc.duqueco@conaf.cl"),
    (19, "Liceo Nueva Zelandia",                           10, "Concepción",   "Santa Juana",    "enc.nuevazelandia@conaf.cl"),
    (20, "Vivero Curacautín",                              11, "Malleco",      "Curacautín",     "enc.curacautin@conaf.cl"),
    (21, "Vivero Imperial",                                11, "Cautín",       "Nueva Imperial", "enc.imperial@conaf.cl"),
    (22, "Vivero Huillilemu",                              12, "Valdivia",     "Mariquina",      "enc.huillilemu@conaf.cl"),
    (23, "Vivero San Juan",                                12, "Valdivia",     "Panguipulli",    "enc.sanjuan@conaf.cl"),
    (24, "Vivero Chabranco",                               12, "Ranco",        "Futrono",        "enc.chabranco@conaf.cl"),
    (25, "Vivero Liceo Técnico de Ignao",                  12, "Ranco",        "Lago Ranco",     "enc.ignao@conaf.cl"),
    (26, "Vivero Mashue",                                  12, "Ranco",        "La Unión",       "enc.mashue@conaf.cl"),
    (27, "Vivero Llancacura",                              12, "Ranco",        "La Unión",       "enc.llancacura@conaf.cl"),
    (28, "Vivero Butalcura",                               13, "Chiloé",       "Dalcahue",       "enc.butalcura@conaf.cl"),
    (29, "Vivero Augusto Falcón",                          14, "General Carrera", "Chile Chico", "enc.augustofalcon@conaf.cl"),
    (30, "Centro de Acondicionamiento Cochrane",           14, "Capitán Prat", "Cochrane",       "enc.cochrane@conaf.cl"),
    (31, "Vivero El Mallín",                               14, "Aysén",        "Aysén",          "enc.mallin@conaf.cl"),
    (32, "Vivero Las Lengas",                              14, "Coyhaique",    "Coyhaique",       "enc.laslengas@conaf.cl"),
    (33, "Vivero Dorotea",                                 15, "Última Esperanza", "Natales",     "enc.dorotea@conaf.cl"),
    (34, "Vivero Río de Los Ciervos",                      15, "Magallanes",   "Punta Arenas",   "enc.ciervos@conaf.cl"),
    (35, "Vivero Buin",                                    16, "Maipo",        "Buin",           "enc.buin@conaf.cl"),
    (36, "Vivero Reserva Nacional Río Clarillo",           16, "Cordillera",   "Pirque",         "enc.clarillo@conaf.cl"),
    (37, "Vivero San Pedro",                               16, "Melipilla",    "San Pedro",      "enc.sanpedro@conaf.cl"),
    (38, "Vivero Mataveri Otai",                           17, "Isla de Pascua", "Isla de Pascua", "enc.matavari@conaf.cl"),
]

# (especie_id, nombre_cientifico, nombre_comun, origen)
ESPECIES = [
    (1,  "Quillaja saponaria",       "Quillay",                 "Nativo"),
    (2,  "Cryptocarya alba",         "Peumo",                   "Nativo"),
    (3,  "Lithraea caustica",        "Litre",                   "Nativo"),
    (4,  "Schinus molle",            "Pimiento",                "Nativo"),
    (5,  "Schinus latifolius",       "Molle",                   "Nativo"),
    (6,  "Maytenus boaria",          "Maitén",                  "Nativo"),
    (7,  "Acacia caven",             "Espino",                  "Nativo"),
    (8,  "Azara dentata",            "Corcolén",                "Nativo"),
    (9,  "Citronella mucronata",     "Naranjillo",              "Nativo"),
    (10, "Crinodendron patagua",     "Patagua",                 "Nativo"),
    (11, "Beilschmiedia miersii",    "Belloto del norte",       "Nativo Conservación"),
    (12, "Beilschmiedia berteroana", "Belloto del sur",         "Nativo Conservación"),
    (13, "Nothofagus obliqua",       "Roble",                   "Nativo"),
    (14, "Nothofagus alessandrii",   "Ruil",                    "Nativo Conservación"),
    (15, "Nothofagus dombeyi",       "Coihue",                  "Nativo"),
    (16, "Nothofagus pumilio",       "Lenga",                   "Nativo"),
    (17, "Nothofagus antarctica",    "Ñirre",                   "Nativo"),
    (18, "Persea lingue",            "Lingue",                  "Nativo"),
    (19, "Aristotelia chilensis",    "Maqui",                   "Nativo"),
    (20, "Drimys winteri",           "Canelo",                  "Nativo Conservación"),
    (21, "Araucaria araucana",       "Araucaria",               "Nativo Conservación"),
    (22, "Austrocedrus chilensis",   "Ciprés de la Cordillera", "Nativo Conservación"),
    (23, "Aextoxicon punctatum",     "Olivillo",                "Nativo Conservación"),
    (24, "Luma apiculata",           "Arrayán",                 "Nativo"),
    (25, "Amomyrtus luma",           "Luma",                    "Nativo"),
    (26, "Amomyrtus meli",           "Meli",                    "Nativo"),
    (27, "Gevuina avellana",         "Avellano chileno",        "Nativo"),
    (28, "Embothrium coccineum",     "Notro",                   "Nativo"),
    (29, "Berberis microphylla",     "Calafate",                "Nativo"),
    (30, "Prosopis chilensis",       "Algarrobo chileno",       "Nativo"),
    (31, "Geoffroea decorticans",    "Chañar",                  "Nativo"),
    (32, "Caesalpinia spinosa",      "Tara",                    "Nativo"),
    (33, "Atriplex atacamensis",     "Cachiyuyo",               "Nativo"),
    (34, "Adesmia confusa",          "Palhuén",                 "Nativo"),
    (35, "Balsamocarpon brevifolium","Algarrobilla",            "Nativo"),
    (36, "Azorella compacta",        "Yareta",                  "Nativo Conservación"),
    (37, "Avellanita bustillosii",   "Avellanita",              "Nativo Conservación"),
    (38, "Sophora cassioides",       "Pelú",                    "Nativo"),
    (39, "Sophora toromiro",         "Toromiro",                "Nativo Conservación"),
    # Especies adicionales usadas por COIPO_INVENTARIO_PLANTA (catálogo = superset).
    (40, "Peumus boldus",            "Boldo",                   "Nativo"),
    (41, "Prosopis tamarugo",        "Tamarugo",                "Nativo Conservación"),
]


def codigo_especie(nombre_comun):
    return "ESP-" + slug(nombre_comun)

# (vivero_id, especie_id, cantidad_disponible)
STOCK = [
    (1, 30, 180), (1, 31, 220), (1, 33, 340), (1, 32, 90),
    (2, 33, 120), (2, 36, 45), (2, 34, 80),
    (3, 30, 260), (3, 31, 310), (3, 33, 420), (3, 32, 75),
    (4, 30, 140), (4, 33, 230), (4, 34, 95),
    (5, 30, 80), (5, 33, 150),
    (6, 31, 280), (6, 33, 200), (6, 34, 120), (6, 35, 95),
    (7, 31, 240), (7, 32, 110), (7, 35, 150),
    (8, 1, 320), (8, 7, 450), (8, 30, 180), (8, 35, 100),
    (9, 23, 70), (9, 8, 50),
    (10, 1, 580), (10, 2, 410), (10, 3, 350), (10, 6, 290), (10, 7, 220), (10, 19, 180),
    (11, 1, 240), (11, 6, 180), (11, 7, 300), (11, 11, 60),
    (12, 1, 320), (12, 2, 260), (12, 7, 420), (12, 11, 40),
    (13, 1, 800), (13, 2, 540), (13, 3, 380), (13, 6, 420), (13, 13, 670), (13, 19, 260),
    (14, 13, 540), (14, 14, 120), (14, 18, 280), (14, 19, 410),
    (15, 13, 320), (15, 14, 80), (15, 24, 150),
    (35, 1, 250), (35, 2, 180), (35, 6, 150), (35, 7, 410), (35, 8, 90), (35, 19, 320),
    (36, 1, 120), (36, 2, 95), (36, 3, 200), (36, 6, 60), (36, 10, 45), (36, 37, 25),
    (37, 1, 180), (37, 7, 220), (37, 9, 80),
    (16, 1, 1200), (16, 13, 1500), (16, 19, 600), (16, 20, 320), (16, 21, 180), (16, 27, 240),
    (17, 13, 280), (17, 24, 160), (17, 25, 140),
    (18, 13, 540), (18, 15, 380), (18, 18, 260), (18, 19, 410),
    (19, 13, 220), (19, 25, 180), (19, 27, 90),
    (20, 13, 1200), (20, 15, 850), (20, 18, 800), (20, 20, 350), (20, 21, 90), (20, 22, 120), (20, 27, 220),
    (21, 13, 680), (21, 15, 420), (21, 24, 380), (21, 27, 180),
    (22, 13, 950), (22, 15, 720), (22, 18, 470), (22, 20, 280), (22, 23, 220), (22, 25, 310),
    (23, 13, 540), (23, 20, 230), (23, 24, 280),
    (24, 13, 420), (24, 15, 380), (24, 25, 290),
    (25, 13, 180), (25, 24, 150),
    (26, 13, 320), (26, 18, 240), (26, 25, 200),
    (27, 13, 280), (27, 20, 180), (27, 23, 140),
    (28, 15, 480), (28, 20, 350), (28, 23, 280), (28, 24, 320), (28, 28, 190),
    (29, 16, 420), (29, 17, 380), (29, 28, 220), (29, 29, 310),
    (30, 16, 280), (30, 17, 240),
    (31, 15, 540), (31, 16, 480), (31, 17, 420), (31, 28, 180),
    (32, 16, 850), (32, 17, 720), (32, 28, 320), (32, 29, 280),
    (33, 16, 380), (33, 17, 320), (33, 29, 240),
    (34, 16, 620), (34, 17, 540), (34, 28, 290), (34, 29, 380),
    (38, 39, 45), (38, 38, 30),
]

# Parámetros por defecto por región (los valores globales actuales del sistema).
PARAM_DEFAULTS = dict(
    plazo_retiro_dias=30,
    max_especies=3,
    max_unidades=100,
    max_solicitudes_anio=1,
)


def slug(texto):
    s = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").upper()
    return s


def codigo_vivero(nombre):
    s = slug(nombre)
    for pref in ("VIVERO-", "CENTRO-DE-ACOPIO-", "CENTRO-EXPERIMENTAL-",
                 "CENTRO-DE-ACONDICIONAMIENTO-", "BASE-BRIGADA-", "LICEO-"):
        if s.startswith(pref):
            s = s[len(pref):]
            break
    return "VIV-" + s


# ---------------------------------------------------------------------------
# SEED — escribe los .xlsx desde los datos canónicos
# ---------------------------------------------------------------------------

def _write_sheet(path, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "datos"
    ws.append(header)
    for r in rows:
        ws.append(list(r))
    wb.save(path)


def seed():
    rcv_rows = []
    region_nombre = {rid: nom for rid, nom in REGIONES}
    for vid, nombre, rid, provincia, comuna, email in VIVEROS:
        rcv_rows.append([
            rid, region_nombre[rid], provincia, comuna,
            vid, nombre, codigo_vivero(nombre), email,
        ])
    _write_sheet(
        XLSX["region_comuna_vivero"],
        ["region_id", "region", "provincia", "comuna", "vivero_id", "vivero", "codigo", "encargado_email"],
        rcv_rows,
    )

    _write_sheet(
        XLSX["plantas"],
        ["especie_id", "nombre_comun", "nombre_cientifico", "origen", "codigo"],
        [[eid, comun, cient, origen, codigo_especie(comun)] for eid, cient, comun, origen in ESPECIES],
    )

    _write_sheet(
        XLSX["stock"],
        ["vivero_id", "especie_id", "cantidad_disponible"],
        [list(r) for r in STOCK],
    )

    _write_sheet(
        XLSX["parametros_region"],
        ["region_id", "plazo_retiro_dias", "max_especies", "max_unidades", "max_solicitudes_anio"],
        [[rid,
          PARAM_DEFAULTS["plazo_retiro_dias"],
          PARAM_DEFAULTS["max_especies"],
          PARAM_DEFAULTS["max_unidades"],
          PARAM_DEFAULTS["max_solicitudes_anio"]] for rid, _ in REGIONES],
    )
    print("[seed] .xlsx escritos en", BASE)


# ---------------------------------------------------------------------------
# BUILD — lee los .xlsx y emite catalogo.generated.json
# ---------------------------------------------------------------------------

def _read_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["datos"] if "datos" in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(it)]
    rows = []
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        rows.append(dict(zip(header, row)))
    return rows


def _i(v):
    return int(v) if v is not None and str(v).strip() != "" else None


def build():
    rcv = _read_rows(XLSX["region_comuna_vivero"])
    plantas = _read_rows(XLSX["plantas"])
    stock = _read_rows(XLSX["stock"])
    params = _read_rows(XLSX["parametros_region"])

    regiones = []
    vistas = set()
    for r in rcv:
        rid = _i(r["region_id"])
        if rid not in vistas:
            vistas.add(rid)
            regiones.append({"id": rid, "nombre": r["region"]})
    regiones.sort(key=lambda x: x["id"])

    viveros = [{
        "id": _i(r["vivero_id"]),
        "nombre": r["vivero"],
        "regionId": _i(r["region_id"]),
        "region": r["region"],
        "provincia": r.get("provincia") or "",
        "comuna": r["comuna"],
        "codigo": r.get("codigo") or "",
        "encargadoEmail": r.get("encargado_email") or "",
    } for r in rcv]
    viveros.sort(key=lambda x: x["id"])

    especies = [{
        "id": _i(p["especie_id"]),
        "nombreCientifico": p["nombre_cientifico"],
        "nombreComun": p["nombre_comun"],
        "origen": p["origen"],
        "codigo": p.get("codigo") or codigo_especie(p["nombre_comun"]),
    } for p in plantas]
    especies.sort(key=lambda x: x["id"])

    stock_json = [{
        "viveroId": _i(s["vivero_id"]),
        "especieId": _i(s["especie_id"]),
        "cantidadDisponible": _i(s["cantidad_disponible"]),
    } for s in stock]

    parametros = [{
        "regionId": _i(p["region_id"]),
        "plazoRetiroDias": _i(p["plazo_retiro_dias"]),
        "maxEspecies": _i(p["max_especies"]),
        "maxUnidades": _i(p["max_unidades"]),
        "maxSolicitudesAnio": _i(p["max_solicitudes_anio"]),
    } for p in params]
    parametros.sort(key=lambda x: x["regionId"])

    catalogo = {
        "_meta": {
            "fuente": "tabla_insumo/*.xlsx",
            "generador": "tabla_insumo/generar_insumos.py",
            "nota": "NO editar a mano: regenerar con `python generar_insumos.py`.",
        },
        "regiones": regiones,
        "viveros": viveros,
        "especies": especies,
        "stock": stock_json,
        "parametrosRegion": parametros,
    }

    payload = json.dumps(catalogo, ensure_ascii=False, indent=2)
    for target in JSON_TARGETS:
        os.makedirs(os.path.dirname(target), exist_ok=True)
        with open(target, "w", encoding="utf-8") as f:
            f.write(payload + "\n")
        print("[build] catalogo ->", target)

    print("[build] regiones=%d viveros=%d especies=%d stock=%d parametros=%d" % (
        len(regiones), len(viveros), len(especies), len(stock_json), len(parametros)))


def main():
    args = set(sys.argv[1:])
    falta = any(not os.path.exists(p) for p in XLSX.values())
    if "--seed" in args or (falta and "--build" not in args):
        seed()
    build()


if __name__ == "__main__":
    main()
