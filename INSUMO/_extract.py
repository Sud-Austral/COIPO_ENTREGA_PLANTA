import sys
import os
import pypdf
import openpyxl

base = os.path.dirname(os.path.abspath(__file__))

pdfs = [
    "Documento Adjunto 1. Instructivo Reportes del SIGI GBCC 8 2026.pdf",
    "Documento Adjunto 2. Ejemplo de Resolución de Baja de Plantas (Resolución 289-2022 La Araucanía)..pdf",
]
xlsxs = [
    "Documento Adjunto 3. SIGI GBCC 8 2026 Programa Produccion y Evaluacion de Plantas.xlsx",
    "Documento Adjunto 4. SIGI GBCC 8 2026 Control de Inventario de Plantas Ingresos y Egresos.xlsx",
]

for p in pdfs:
    path = os.path.join(base, p)
    out = os.path.join(base, "_" + os.path.splitext(p)[0] + ".txt")
    try:
        r = pypdf.PdfReader(path)
        with open(out, "w", encoding="utf-8") as f:
            f.write(f"=== {p} (pages: {len(r.pages)}) ===\n\n")
            for i, page in enumerate(r.pages):
                f.write(f"\n--- Page {i+1} ---\n")
                try:
                    f.write(page.extract_text() or "")
                except Exception as e:
                    f.write(f"[ERROR extracting page {i+1}: {e}]")
        print(f"Wrote {out}")
    except Exception as e:
        print(f"FAILED {p}: {e}")

for x in xlsxs:
    path = os.path.join(base, x)
    out = os.path.join(base, "_" + os.path.splitext(x)[0] + ".txt")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        with open(out, "w", encoding="utf-8") as f:
            f.write(f"=== {x} ===\n")
            f.write(f"Sheets: {wb.sheetnames}\n\n")
            for sn in wb.sheetnames:
                ws = wb[sn]
                f.write(f"\n--- Sheet: {sn} ---\n")
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None and str(c).strip() != "" for c in row):
                        f.write(" | ".join("" if c is None else str(c) for c in row) + "\n")
        print(f"Wrote {out}")
    except Exception as e:
        print(f"FAILED {x}: {e}")
